import math
import os
from io import BytesIO
from pathlib import Path

from PyQt5.QtCore import QThread, pyqtSignal
from PIL import Image as PILImage
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.styles import Font as XLFont, Alignment as XLAlignment, Border, Side, PatternFill

from app.core.models import CM_TO_PX_96


# ── Worker thread ──────────────────────────────────────────────────────────────
class InsertWorker(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal(str)
    status = pyqtSignal(str)

    def __init__(self, params):
        super().__init__()
        self.p = params

    def run(self):
        try:
            self._do_insert()
            self.finished.emit("")
        except Exception as e:
            self.finished.emit(str(e))

    @staticmethod
    def _col_width_px(ws, col_idx):
        letter = get_column_letter(col_idx)
        w = ws.column_dimensions[letter].width
        if w is None:
            w = 8.43
        return w * 7 + 5

    @staticmethod
    def _row_height_px(ws, row_idx):
        h = ws.row_dimensions[row_idx].height
        if h is None:
            h = 15
        return h * 4 / 3

    def _do_insert(self):
        p = self.p
        if p["excel_path"] and os.path.exists(p["excel_path"]):
            wb = openpyxl.load_workbook(p["excel_path"])
        else:
            wb = openpyxl.Workbook()
            # Remove default "Sheet" if creating new
            if "Sheet" in wb.sheetnames and p.get("sheet_name") != "Sheet":
                del wb["Sheet"]

        # Create or get the target sheet
        sheet_name = p["sheet_name"]
        if p["sheet_new"]:
            ws = wb.create_sheet(title=sheet_name)
            # Move sheet to requested position
            insert_after_name = p.get("insert_after_name", None)
            if insert_after_name and insert_after_name in wb.sheetnames:
                target_idx = wb.sheetnames.index(insert_after_name) + 1
                current_idx = len(wb.sheetnames) - 1
                wb.move_sheet(ws, offset=target_idx - current_idx)
        else:
            ws = wb[sheet_name]

        groups = p["groups"]
        cols = p["grid_cols"]
        start_col_idx = openpyxl.utils.column_index_from_string(p["start_col"])
        start_row = p["start_row"]
        use_groups = p.get("use_groups", False)

        total_images = sum(len(g["images"]) for g in groups)
        processed = 0
        current_row = start_row

        # Reserve rows for inline sheet TOC (collapsible group list)
        inline_toc = use_groups and p.get("create_toc", False)
        inline_toc_start = None
        if inline_toc:
            # Row for "Contents" label
            col_letter = get_column_letter(start_col_idx)
            ws[f"{col_letter}{current_row}"] = "▸ Contents"
            ws[f"{col_letter}{current_row}"].font = XLFont(bold=True, size=11, color="1F4E79")
            current_row += 1
            inline_toc_start = current_row
            # Reserve one row per group (will fill with links after placing images)
            for _ in groups:
                current_row += 1
            current_row += 1  # blank row before images

        toc_entries = []
        group_header_rows = []  # track header row for each group

        for group in groups:
            title = group["title"]
            images = group["images"]

            if use_groups:
                header_cell = f"{get_column_letter(start_col_idx)}{current_row}"
                ws[header_cell] = title
                ws[header_cell].font = XLFont(bold=True, size=12)
                ws[header_cell].alignment = XLAlignment(vertical="center")
                ws.row_dimensions[current_row].height = 22
                toc_entries.append((title, sheet_name, header_cell))
                group_header_rows.append(current_row)
                current_row += 1

            for i, img_path in enumerate(images):
                self.status.emit(f"Processing {processed+1}/{total_images}: {Path(img_path).name}")

                img = PILImage.open(img_path).convert("RGB")

                if p["crop_ratio"]:
                    img = self._crop_center(img, p["crop_ratio"])
                if p["resize_px_w"] or p["resize_px_h"]:
                    img = self._resize_px(img, p["resize_px_w"], p["resize_px_h"])

                buf = BytesIO()
                img.save(buf, format="JPEG", quality=90)
                buf.seek(0)

                xl_img = XLImage(buf)
                w_cm = p["display_w_cm"]
                h_cm = p["display_h_cm"]
                display_mode = p.get("display_mode", 1)
                anchor_axis = p.get("anchor_axis", "W")
                fixed_aspect = p.get("fixed_aspect", (4, 3))
                if display_mode == 0:  # Per image — derive from image's own aspect
                    iw, ih = img.size
                    if iw > 0 and ih > 0:
                        if anchor_axis == "W":
                            h_cm = w_cm * (ih / iw)
                        else:  # H
                            w_cm = h_cm * (iw / ih)
                elif display_mode == 1:  # Fixed ratio — derive from fixed_aspect
                    aw, ah = fixed_aspect
                    if aw > 0 and ah > 0:
                        if anchor_axis == "W":
                            h_cm = w_cm * (ah / aw)
                        else:
                            w_cm = h_cm * (aw / ah)
                # else: Manual — use w_cm and h_cm as given
                xl_img.width = w_cm * CM_TO_PX_96
                xl_img.height = h_cm * CM_TO_PX_96

                row_offset = i // cols
                col_offset = i % cols
                img_w_px = xl_img.width
                img_h_px = xl_img.height

                if p["placement"] == "in_cell":
                    cell_col = start_col_idx + col_offset
                    cell_row = current_row + row_offset
                    ws.column_dimensions[get_column_letter(cell_col)].width = w_cm * 4.8
                    ws.row_dimensions[cell_row].height = h_cm * 28.35
                    ws.add_image(xl_img, f"{get_column_letter(cell_col)}{cell_row}")
                else:
                    gap_h_px = p.get("gap_h_cm", 0.5) * CM_TO_PX_96
                    gap_v_px = p.get("gap_v_cm", 0.5) * CM_TO_PX_96
                    # Pure EMU offsets from the grid's top-left anchor cell.
                    # The X/Y position of this image relative to the start cell
                    # is the SUM of all preceding image widths and gaps; we
                    # encode it entirely in colOff/rowOff so the actual cell
                    # widths/heights of the sheet never enter the calculation
                    # (which is what caused the horizontal-collapse /
                    # vertical-stretch artefacts on default-sized sheets).
                    x_px = col_offset * (img_w_px + gap_h_px)
                    y_px = row_offset * (img_h_px + gap_v_px)
                    emu_w = pixels_to_EMU(img_w_px)
                    emu_h = pixels_to_EMU(img_h_px)

                    marker = AnchorMarker(
                        col=start_col_idx - 1,
                        colOff=pixels_to_EMU(x_px),
                        row=current_row - 1,
                        rowOff=pixels_to_EMU(y_px),
                    )
                    anchor = OneCellAnchor(
                        _from=marker,
                        ext=XDRPositiveSize2D(cx=emu_w, cy=emu_h),
                    )
                    xl_img.anchor = anchor
                    ws.add_image(xl_img)

                processed += 1
                self.progress.emit(int(processed / total_images * 100))

            image_rows = math.ceil(len(images) / cols) if images else 0
            if p["placement"] == "in_cell":
                current_row += image_rows
            else:
                img_total_h_px = image_rows * (h_cm * CM_TO_PX_96 + p.get("gap_v_cm", 0.5) * CM_TO_PX_96)
                rows_consumed = 1
                h_acc = 0
                while h_acc < img_total_h_px:
                    h_acc += self._row_height_px(ws, current_row + rows_consumed - 1)
                    rows_consumed += 1
                current_row += rows_consumed

            if use_groups:
                current_row += 1

        # ── Inline sheet TOC (collapsible) ───────────────────────────────
        if inline_toc and group_header_rows:
            col_letter = get_column_letter(start_col_idx)
            for gi, (group, header_row) in enumerate(zip(groups, group_header_rows)):
                toc_r = inline_toc_start + gi
                cell = f"{col_letter}{toc_r}"
                ws[cell] = f"    {group['title']}"
                ws[cell].font = XLFont(size=10, color="0563C1", underline="single")
                ws[cell].hyperlink = f"#'{sheet_name}'!{col_letter}{header_row}"
            # Group and collapse the TOC rows
            ws.row_dimensions.group(inline_toc_start, inline_toc_start + len(groups) - 1,
                                    hidden=True, outline_level=1)

        # ── TOC sheet ─────────────────────────────────────────────────────
        if p.get("create_toc", False) and toc_entries:
            toc_name = "Contents"
            thin_border = Border(
                left=Side(style="thin", color="D0D0D0"),
                right=Side(style="thin", color="D0D0D0"),
                top=Side(style="thin", color="D0D0D0"),
                bottom=Side(style="thin", color="D0D0D0"),
            )
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            sheet_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

            toc_existed = toc_name in wb.sheetnames
            if toc_existed:
                toc_ws = wb[toc_name]
            else:
                toc_ws = wb.create_sheet(title=toc_name, index=0)

            # Collect existing TOC entries from other sheets (by scanning TOC rows)
            existing_sections = []  # [(sheet_name, [(title, cell_ref)])]
            if toc_existed:
                r = 2
                max_r = toc_ws.max_row
                while r <= max_r:
                    cell_val = toc_ws[f"A{r}"].value
                    if cell_val and str(cell_val).startswith("▸"):
                        sec_name = str(cell_val)[2:].strip()
                        sec_entries = []
                        r += 1
                        while r <= max_r:
                            b_val = toc_ws[f"B{r}"].value
                            if not b_val:
                                r += 1
                                break
                            a_val = toc_ws[f"A{r}"].value
                            if a_val and str(a_val).startswith("▸"):
                                break
                            link = toc_ws[f"B{r}"].hyperlink
                            href = link.target if link else f"#'{sec_name}'!A1"
                            sec_entries.append((str(b_val), sec_name, href))
                            r += 1
                        existing_sections.append((sec_name, sec_entries))
                    else:
                        r += 1

            # Replace or add the current sheet's section
            new_section = (sheet_name, [(t, sn, f"#'{sn}'!{cr}") for t, sn, cr in toc_entries])
            replaced = False
            for i, (sn, _) in enumerate(existing_sections):
                if sn == sheet_name:
                    existing_sections[i] = new_section
                    replaced = True
                    break
            if not replaced:
                existing_sections.append(new_section)

            # Sort sections by workbook sheet order
            sheet_order = {name: idx for idx, name in enumerate(wb.sheetnames)}
            existing_sections.sort(key=lambda s: sheet_order.get(s[0], 999))

            # Clear and rewrite entire TOC
            for row in toc_ws.iter_rows(min_row=1, max_row=toc_ws.max_row, max_col=3):
                for cell in row:
                    cell.value = None
                    cell.font = XLFont()
                    cell.fill = PatternFill()
                    cell.border = Border()
                    cell.hyperlink = None
                    cell.alignment = XLAlignment()

            # Header
            toc_ws["A1"] = "Contents"
            toc_ws["A1"].font = XLFont(bold=True, size=16, color="FFFFFF")
            toc_ws["A1"].fill = header_fill
            toc_ws["A1"].alignment = XLAlignment(vertical="center")
            toc_ws["B1"].fill = header_fill
            toc_ws["C1"].fill = header_fill
            toc_ws.row_dimensions[1].height = 32
            toc_ws.column_dimensions["A"].width = 6
            toc_ws.column_dimensions["B"].width = 40
            toc_ws.column_dimensions["C"].width = 15

            toc_row = 3
            for sec_name, sec_entries in existing_sections:
                toc_ws[f"A{toc_row}"] = f"▸ {sec_name}"
                toc_ws[f"A{toc_row}"].font = XLFont(bold=True, size=11, color="1F4E79")
                toc_ws[f"A{toc_row}"].fill = sheet_fill
                toc_ws[f"B{toc_row}"].fill = sheet_fill
                toc_ws[f"C{toc_row}"].fill = sheet_fill
                toc_ws[f"A{toc_row}"].hyperlink = f"#'{sec_name}'!A1"
                toc_ws[f"A{toc_row}"].border = thin_border
                toc_ws[f"B{toc_row}"].border = thin_border
                toc_ws.row_dimensions[toc_row].height = 22
                toc_row += 1
                for title, sn, href in sec_entries:
                    toc_ws[f"B{toc_row}"] = title
                    toc_ws[f"B{toc_row}"].font = XLFont(size=10, color="0563C1", underline="single")
                    toc_ws[f"B{toc_row}"].hyperlink = href
                    toc_ws[f"B{toc_row}"].border = thin_border
                    toc_ws[f"A{toc_row}"].border = thin_border
                    toc_row += 1
                toc_row += 1  # blank row between sections

        save_path = p["save_path"] or p["excel_path"]
        self.status.emit(f"Saving {save_path}...")
        wb.save(save_path)

    @staticmethod
    def _crop_center(img, ratio):
        w, h = img.size
        target_aspect = ratio[0] / ratio[1]
        current_aspect = w / h
        if current_aspect > target_aspect:
            new_w = int(h * target_aspect)
            left = (w - new_w) // 2
            return img.crop((left, 0, left + new_w, h))
        else:
            new_h = int(w / target_aspect)
            top = (h - new_h) // 2
            return img.crop((0, top, w, top + new_h))

    @staticmethod
    def _resize_px(img, max_w, max_h):
        w, h = img.size
        if max_w and max_h:
            ratio = min(max_w / w, max_h / h)
        elif max_w:
            ratio = max_w / w
        else:
            ratio = max_h / h
        if ratio < 1:
            img = img.resize((int(w * ratio), int(h * ratio)), PILImage.LANCZOS)
        return img
