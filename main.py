#!/usr/bin/env python3
"""Excel Image Inserter — PyQt5 utility for batch-inserting images into Excel."""

import sys
import os
import math
from pathlib import Path
from io import BytesIO

from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QGroupBox, QLabel, QPushButton, QComboBox, QSpinBox, QDoubleSpinBox,
    QLineEdit, QFileDialog, QListWidget, QListWidgetItem, QAbstractItemView,
    QRadioButton, QButtonGroup, QMessageBox, QProgressBar, QCheckBox,
    QFrame, QGridLayout, QSizePolicy, QScrollArea, QToolTip,
    QTreeWidget, QTreeWidgetItem, QHeaderView, QProgressDialog,
    QInputDialog, QMenu,
)
from PyQt5.QtCore import Qt, QSize, QThread, pyqtSignal, QRect, QPoint, QTimer
from PyQt5.QtGui import QPixmap, QIcon, QImage, QPainter, QPen, QColor, QFont, QBrush

from PIL import Image as PILImage
import openpyxl
from openpyxl.drawing.image import Image as XLImage
from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor, AnchorMarker
from openpyxl.drawing.xdr import XDRPositiveSize2D
from openpyxl.utils import get_column_letter
from openpyxl.utils.units import pixels_to_EMU
from openpyxl.styles import Font as XLFont, Alignment as XLAlignment, Border, Side, PatternFill


# ── Constants ──────────────────────────────────────────────────────────────────
APP_VERSION = "1.3.0"
BUILD_NUMBER = "20260417"
CM_TO_PX_96 = 96 / 2.54
CROP_PRESETS = {
    "None": None,
    "1:1": (1, 1),
    "4:3": (4, 3),
    "3:2": (3, 2),
    "16:9": (16, 9),
    "3:4": (3, 4),
    "2:3": (2, 3),
    "9:16": (9, 16),
}
THUMB_SIZE = 100
GROUP_ICON = "\u25bc"
GROUP_ICON_COLLAPSED = "\u25b6"


def estimate_size(path, max_w, max_h):
    """Return (original_mb, estimated_mb, width, height)."""
    size_mb = os.path.getsize(path) / (1024 * 1024)
    try:
        img = PILImage.open(path)
        w, h = img.size
        if max_w or max_h:
            ratio = 1.0
            if max_w and max_h:
                ratio = min(max_w / w, max_h / h)
            elif max_w:
                ratio = max_w / w
            else:
                ratio = max_h / h
            if ratio < 1:
                new_pixels = int(w * ratio) * int(h * ratio)
            else:
                new_pixels = w * h
            est_mb = new_pixels * 0.5 / (1024 * 1024)
        else:
            est_mb = size_mb
        return size_mb, est_mb, w, h
    except Exception:
        return size_mb, size_mb, 0, 0


# ── Worker thread ──────────────────────────────────────────────────────────────
class InsertWorker(QThread):
    progress = pyqtSignal(int)
    finished = pyqtSignal(str)
    status = pyqtSignal(str)

    def __init__(self, params):
        super().__init__()
        self.p = params

    def run(self):
        try:
            self._do_insert()
            self.finished.emit("")
        except Exception as e:
            self.finished.emit(str(e))

    @staticmethod
    def _col_width_px(ws, col_idx):
        letter = get_column_letter(col_idx)
        w = ws.column_dimensions[letter].width
        if w is None:
            w = 8.43
        return w * 7 + 5

    @staticmethod
    def _row_height_px(ws, row_idx):
        h = ws.row_dimensions[row_idx].height
        if h is None:
            h = 15
        return h * 4 / 3

    def _do_insert(self):
        p = self.p
        if p["excel_path"] and os.path.exists(p["excel_path"]):
            wb = openpyxl.load_workbook(p["excel_path"])
        else:
            wb = openpyxl.Workbook()
            # Remove default "Sheet" if creating new
            if "Sheet" in wb.sheetnames and p.get("sheet_name") != "Sheet":
                del wb["Sheet"]

        # Create or get the target sheet
        sheet_name = p["sheet_name"]
        if p["sheet_new"]:
            ws = wb.create_sheet(title=sheet_name)
            # Move sheet to requested position
            insert_after_name = p.get("insert_after_name", None)
            if insert_after_name and insert_after_name in wb.sheetnames:
                target_idx = wb.sheetnames.index(insert_after_name) + 1
                current_idx = len(wb.sheetnames) - 1
                wb.move_sheet(ws, offset=target_idx - current_idx)
        else:
            ws = wb[sheet_name]

        groups = p["groups"]
        cols = p["grid_cols"]
        start_col_idx = openpyxl.utils.column_index_from_string(p["start_col"])
        start_row = p["start_row"]
        use_groups = p.get("use_groups", False)

        total_images = sum(len(g["images"]) for g in groups)
        processed = 0
        current_row = start_row

        toc_entries = []

        for group in groups:
            title = group["title"]
            images = group["images"]

            if use_groups:
                header_cell = f"{get_column_letter(start_col_idx)}{current_row}"
                ws[header_cell] = title
                ws[header_cell].font = XLFont(bold=True, size=12)
                ws[header_cell].alignment = XLAlignment(vertical="center")
                ws.row_dimensions[current_row].height = 22
                toc_entries.append((title, sheet_name, header_cell))
                current_row += 1

            for i, img_path in enumerate(images):
                self.status.emit(f"Processing {processed+1}/{total_images}: {Path(img_path).name}")

                img = PILImage.open(img_path).convert("RGB")

                if p["crop_ratio"]:
                    img = self._crop_center(img, p["crop_ratio"])
                if p["resize_px_w"] or p["resize_px_h"]:
                    img = self._resize_px(img, p["resize_px_w"], p["resize_px_h"])

                buf = BytesIO()
                img.save(buf, format="JPEG", quality=90)
                buf.seek(0)

                xl_img = XLImage(buf)
                w_cm = p["display_w_cm"]
                h_cm = p["display_h_cm"]
                display_mode = p.get("display_mode", 2)
                if display_mode == 0:
                    iw, ih = img.size
                    if iw > 0 and ih > 0:
                        aspect = iw / ih
                        if aspect >= 1:
                            h_cm = w_cm / aspect
                        else:
                            w_cm = h_cm * aspect
                xl_img.width = w_cm * CM_TO_PX_96
                xl_img.height = h_cm * CM_TO_PX_96

                row_offset = i // cols
                col_offset = i % cols
                img_w_px = xl_img.width
                img_h_px = xl_img.height

                if p["placement"] == "in_cell":
                    cell_col = start_col_idx + col_offset
                    cell_row = current_row + row_offset
                    ws.column_dimensions[get_column_letter(cell_col)].width = w_cm * 4.8
                    ws.row_dimensions[cell_row].height = h_cm * 28.35
                    ws.add_image(xl_img, f"{get_column_letter(cell_col)}{cell_row}")
                else:
                    gap_h_px = p.get("gap_h_cm", 0.5) * CM_TO_PX_96
                    gap_v_px = p.get("gap_v_cm", 0.5) * CM_TO_PX_96
                    x_px = col_offset * (img_w_px + gap_h_px)
                    y_px = row_offset * (img_h_px + gap_v_px)
                    emu_w = pixels_to_EMU(img_w_px)
                    emu_h = pixels_to_EMU(img_h_px)

                    col_i = start_col_idx
                    remaining_x = x_px
                    while remaining_x > 0:
                        cw = self._col_width_px(ws, col_i)
                        if remaining_x < cw:
                            break
                        remaining_x -= cw
                        col_i += 1

                    row_i = current_row
                    remaining_y = y_px
                    while remaining_y > 0:
                        rh = self._row_height_px(ws, row_i)
                        if remaining_y < rh:
                            break
                        remaining_y -= rh
                        row_i += 1

                    marker = AnchorMarker(
                        col=col_i - 1,
                        colOff=pixels_to_EMU(remaining_x),
                        row=row_i - 1,
                        rowOff=pixels_to_EMU(remaining_y),
                    )
                    anchor = OneCellAnchor(
                        _from=marker,
                        ext=XDRPositiveSize2D(cx=emu_w, cy=emu_h),
                    )
                    xl_img.anchor = anchor
                    ws.add_image(xl_img)

                processed += 1
                self.progress.emit(int(processed / total_images * 100))

            image_rows = math.ceil(len(images) / cols) if images else 0
            if p["placement"] == "in_cell":
                current_row += image_rows
            else:
                img_total_h_px = image_rows * (h_cm * CM_TO_PX_96 + p.get("gap_v_cm", 0.5) * CM_TO_PX_96)
                rows_consumed = 1
                h_acc = 0
                while h_acc < img_total_h_px:
                    h_acc += self._row_height_px(ws, current_row + rows_consumed - 1)
                    rows_consumed += 1
                current_row += rows_consumed

            if use_groups:
                current_row += 1

        # ── TOC sheet ─────────────────────────────────────────────────────
        if p.get("create_toc", False) and toc_entries:
            toc_name = "Contents"
            thin_border = Border(
                left=Side(style="thin", color="D0D0D0"),
                right=Side(style="thin", color="D0D0D0"),
                top=Side(style="thin", color="D0D0D0"),
                bottom=Side(style="thin", color="D0D0D0"),
            )
            header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
            sheet_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")

            toc_existed = toc_name in wb.sheetnames
            if toc_existed:
                toc_ws = wb[toc_name]
            else:
                toc_ws = wb.create_sheet(title=toc_name, index=0)

            # Collect existing TOC entries from other sheets (by scanning TOC rows)
            existing_sections = []  # [(sheet_name, [(title, cell_ref)])]
            if toc_existed:
                r = 2
                max_r = toc_ws.max_row
                while r <= max_r:
                    cell_val = toc_ws[f"A{r}"].value
                    if cell_val and str(cell_val).startswith("\u25b8"):
                        sec_name = str(cell_val)[2:].strip()
                        sec_entries = []
                        r += 1
                        while r <= max_r:
                            b_val = toc_ws[f"B{r}"].value
                            if not b_val:
                                r += 1
                                break
                            a_val = toc_ws[f"A{r}"].value
                            if a_val and str(a_val).startswith("\u25b8"):
                                break
                            link = toc_ws[f"B{r}"].hyperlink
                            href = link.target if link else f"#'{sec_name}'!A1"
                            sec_entries.append((str(b_val), sec_name, href))
                            r += 1
                        existing_sections.append((sec_name, sec_entries))
                    else:
                        r += 1

            # Replace or add the current sheet's section
            new_section = (sheet_name, [(t, sn, f"#'{sn}'!{cr}") for t, sn, cr in toc_entries])
            replaced = False
            for i, (sn, _) in enumerate(existing_sections):
                if sn == sheet_name:
                    existing_sections[i] = new_section
                    replaced = True
                    break
            if not replaced:
                existing_sections.append(new_section)

            # Sort sections by workbook sheet order
            sheet_order = {name: idx for idx, name in enumerate(wb.sheetnames)}
            existing_sections.sort(key=lambda s: sheet_order.get(s[0], 999))

            # Clear and rewrite entire TOC
            for row in toc_ws.iter_rows(min_row=1, max_row=toc_ws.max_row, max_col=3):
                for cell in row:
                    cell.value = None
                    cell.font = XLFont()
                    cell.fill = PatternFill()
                    cell.border = Border()
                    cell.hyperlink = None
                    cell.alignment = XLAlignment()

            # Header
            toc_ws["A1"] = "Contents"
            toc_ws["A1"].font = XLFont(bold=True, size=16, color="FFFFFF")
            toc_ws["A1"].fill = header_fill
            toc_ws["A1"].alignment = XLAlignment(vertical="center")
            toc_ws["B1"].fill = header_fill
            toc_ws["C1"].fill = header_fill
            toc_ws.row_dimensions[1].height = 32
            toc_ws.column_dimensions["A"].width = 6
            toc_ws.column_dimensions["B"].width = 40
            toc_ws.column_dimensions["C"].width = 15

            toc_row = 3
            for sec_name, sec_entries in existing_sections:
                toc_ws[f"A{toc_row}"] = f"\u25b8 {sec_name}"
                toc_ws[f"A{toc_row}"].font = XLFont(bold=True, size=11, color="1F4E79")
                toc_ws[f"A{toc_row}"].fill = sheet_fill
                toc_ws[f"B{toc_row}"].fill = sheet_fill
                toc_ws[f"C{toc_row}"].fill = sheet_fill
                toc_ws[f"A{toc_row}"].hyperlink = f"#'{sec_name}'!A1"
                toc_ws[f"A{toc_row}"].border = thin_border
                toc_ws[f"B{toc_row}"].border = thin_border
                toc_ws.row_dimensions[toc_row].height = 22
                toc_row += 1
                for title, sn, href in sec_entries:
                    toc_ws[f"B{toc_row}"] = title
                    toc_ws[f"B{toc_row}"].font = XLFont(size=10, color="0563C1", underline="single")
                    toc_ws[f"B{toc_row}"].hyperlink = href
                    toc_ws[f"B{toc_row}"].border = thin_border
                    toc_ws[f"A{toc_row}"].border = thin_border
                    toc_row += 1
                toc_row += 1  # blank row between sections

        save_path = p["save_path"] or p["excel_path"]
        self.status.emit(f"Saving {save_path}...")
        wb.save(save_path)

    @staticmethod
    def _crop_center(img, ratio):
        w, h = img.size
        target_aspect = ratio[0] / ratio[1]
        current_aspect = w / h
        if current_aspect > target_aspect:
            new_w = int(h * target_aspect)
            left = (w - new_w) // 2
            return img.crop((left, 0, left + new_w, h))
        else:
            new_h = int(w / target_aspect)
            top = (h - new_h) // 2
            return img.crop((0, top, w, top + new_h))

    @staticmethod
    def _resize_px(img, max_w, max_h):
        w, h = img.size
        if max_w and max_h:
            ratio = min(max_w / w, max_h / h)
        elif max_w:
            ratio = max_w / w
        else:
            ratio = max_h / h
        if ratio < 1:
            img = img.resize((int(w * ratio), int(h * ratio)), PILImage.LANCZOS)
        return img


# ── Image loader thread ───────────────────────────────────────────────────────
class ImageLoaderThread(QThread):
    progress = pyqtSignal(int, int)
    item_ready = pyqtSignal(str, float, float, int, int)
    finished = pyqtSignal()

    def __init__(self, paths, max_w, max_h):
        super().__init__()
        self.paths = paths
        self.max_w = max_w
        self.max_h = max_h

    def run(self):
        total = len(self.paths)
        for i, p in enumerate(self.paths):
            orig_mb, est_mb, w, h = estimate_size(p, self.max_w, self.max_h)
            self.item_ready.emit(p, orig_mb, est_mb, w, h)
            self.progress.emit(i + 1, total)
        self.finished.emit()


# ── Thumbnail stack widget ─────────────────────────────────────────────────────
class ThumbCard(QWidget):
    delete_requested = pyqtSignal(str)
    selection_toggled = pyqtSignal(str, bool)

    def __init__(self, path, orig_mb, est_mb, w, h):
        super().__init__()
        self.path = path
        self.orig_mb = orig_mb
        self.est_mb = est_mb
        self.img_w = w
        self.img_h = h
        self.selected = False
        self._drag_start = None
        self.pixmap = QPixmap(path).scaled(THUMB_SIZE, THUMB_SIZE, Qt.KeepAspectRatio, Qt.SmoothTransformation)
        self.setFixedSize(self.pixmap.width(), self.pixmap.height())
        self.setToolTip(f"{Path(path).name}\n{w}x{h}\n{orig_mb:.2f} MB -> {est_mb:.2f} MB")

    def paintEvent(self, event):
        p = QPainter(self)
        p.setRenderHint(QPainter.Antialiasing)
        p.drawPixmap(0, 0, self.pixmap)
        if self.selected:
            p.setPen(QPen(QColor("#6366f1"), 3))
            p.setBrush(Qt.NoBrush)
            p.drawRect(1, 1, self.width() - 2, self.height() - 2)
        bar_h = 18
        bar_y = self.height() - bar_h
        p.fillRect(0, bar_y, self.width(), bar_h, QColor(255, 255, 255, 200))
        p.setFont(QFont("Arial", 8))
        p.setPen(QColor("#333"))
        p.drawText(4, bar_y, self.width() // 2, bar_h, Qt.AlignLeft | Qt.AlignVCenter, f"{self.orig_mb:.2f}MB")
        p.setPen(QColor("#16a34a"))
        p.drawText(self.width() // 2, bar_y, self.width() // 2 - 4, bar_h, Qt.AlignRight | Qt.AlignVCenter, f"{self.est_mb:.2f}MB")
        btn_r = 9
        cx = self.width() - btn_r - 4
        cy = btn_r + 4
        p.setBrush(QColor(255, 255, 255, 220))
        p.setPen(Qt.NoPen)
        p.drawEllipse(QPoint(cx, cy), btn_r, btn_r)
        p.setPen(QColor("#333"))
        p.setFont(QFont("Arial", 9, QFont.Bold))
        p.drawText(cx - btn_r, cy - btn_r, btn_r * 2, btn_r * 2, Qt.AlignCenter, "\u00d7")
        p.end()

    def mousePressEvent(self, event):
        btn_r = 9
        cx = self.width() - btn_r - 4
        cy = btn_r + 4
        if (event.pos().x() - cx) ** 2 + (event.pos().y() - cy) ** 2 <= (btn_r + 3) ** 2:
            self.delete_requested.emit(self.path)
            return
        self._drag_start = event.pos()

    def mouseMoveEvent(self, event):
        if self._drag_start and (event.pos() - self._drag_start).manhattanLength() > 10:
            from PyQt5.QtCore import QMimeData
            from PyQt5.QtGui import QDrag
            drag = QDrag(self)
            mime = QMimeData()
            mime.setText(self.path)
            drag.setMimeData(mime)
            drag.setPixmap(self.pixmap.scaled(60, 60, Qt.KeepAspectRatio, Qt.SmoothTransformation))
            drag.exec_(Qt.MoveAction)
            self._drag_start = None

    def mouseReleaseEvent(self, event):
        if self._drag_start:
            self.selected = not self.selected
            self.selection_toggled.emit(self.path, self.selected)
            self.update()
        self._drag_start = None


class ThumbStackView(QScrollArea):
    delete_requested = pyqtSignal(str)
    order_changed = pyqtSignal(list)

    def __init__(self):
        super().__init__()
        self.setWidgetResizable(True)
        self.setAcceptDrops(True)
        self.setStyleSheet("ThumbStackView { border: 1px solid palette(mid); border-radius: 6px; }")
        self.container = QWidget()
        self.container.setAcceptDrops(True)
        self.flow = FlowLayout(self.container)
        self.flow.setSpacing(8)
        self.setWidget(self.container)
        self.cards = []
        self.selected_paths = set()
        self._paths = []

    def set_images(self, paths, max_w, max_h):
        self.flow.clear_widgets()
        for c in self.cards:
            c.setParent(None)
            c.deleteLater()
        self.cards.clear()
        self.selected_paths.clear()
        self._paths = list(paths)
        for path in paths:
            orig_mb, est_mb, w, h = estimate_size(path, max_w, max_h)
            card = ThumbCard(path, orig_mb, est_mb, w, h)
            card.delete_requested.connect(self._on_delete)
            card.selection_toggled.connect(self._on_selection)
            self.cards.append(card)
        self.flow.set_widgets(self.cards)

    def _on_delete(self, path):
        self.delete_requested.emit(path)

    def _on_selection(self, path, selected):
        if selected:
            self.selected_paths.add(path)
        else:
            self.selected_paths.discard(path)

    def get_selected(self):
        return list(self.selected_paths)

    def dragEnterEvent(self, event):
        if event.mimeData().hasText():
            event.acceptProposedAction()

    def dragMoveEvent(self, event):
        event.acceptProposedAction()

    def dropEvent(self, event):
        src_path = event.mimeData().text()
        if src_path not in self._paths:
            return
        drop_pos = self.container.mapFrom(self, event.pos())
        target_idx = len(self._paths) - 1
        for i, card in enumerate(self.cards):
            if card.geometry().contains(drop_pos):
                target_idx = i
                break
        src_idx = self._paths.index(src_path)
        if src_idx == target_idx:
            return
        self._paths.pop(src_idx)
        self._paths.insert(target_idx, src_path)
        self.order_changed.emit(list(self._paths))
        event.acceptProposedAction()


class FlowLayout(QVBoxLayout):
    def __init__(self, parent=None):
        super().__init__(parent)
        self._widgets = []

    def clear_widgets(self):
        self._widgets.clear()
        while self.count():
            item = self.takeAt(0)
            if item.layout():
                while item.layout().count():
                    item.layout().takeAt(0)

    def set_widgets(self, widgets):
        self._widgets = list(widgets)
        self._relayout()

    def addWidget(self, widget):
        self._widgets.append(widget)
        self._relayout()

    def _relayout(self):
        while self.count():
            item = self.takeAt(0)
            if item.layout():
                while item.layout().count():
                    item.layout().takeAt(0)
        if not self._widgets:
            return
        parent = self.parentWidget()
        available_w = parent.width() if parent else 600
        card_w = THUMB_SIZE + 10
        cols = max(1, available_w // card_w)
        row_lay = None
        for i, w in enumerate(self._widgets):
            if i % cols == 0:
                row_lay = QHBoxLayout()
                row_lay.setAlignment(Qt.AlignLeft)
                super().addLayout(row_lay)
            row_lay.addWidget(w)


# ── Grid preview widget ───────────────────────────────────────────────────────
class GridPreview(QWidget):
    HEADER_H = 16
    ROW_NUM_W = 24

    def __init__(self):
        super().__init__()
        self.groups = []
        self.cols = 2
        self.crop_ratio = None
        self.start_col = "A"
        self.start_row = 1
        self.placement = "over"
        self.use_groups = False
        self.setMinimumHeight(100)
        self.setMaximumHeight(160)

    def update_params(self, groups, cols, crop_ratio, start_col="A", start_row=1, placement="over", use_groups=False):
        self.groups = groups
        self.cols = max(1, cols)
        self.crop_ratio = crop_ratio
        self.start_col = start_col
        self.start_row = start_row
        self.placement = placement
        self.use_groups = use_groups
        self.update()

    def paintEvent(self, event):
        painter = QPainter(self)
        painter.setRenderHint(QPainter.Antialiasing)
        total_w = self.width()
        total_h = self.height()
        hh = self.HEADER_H
        rw = self.ROW_NUM_W
        painter.fillRect(self.rect(), QColor("#f0f0f0"))

        total_images = sum(len(g["images"]) for g in self.groups)
        if total_images == 0:
            self._draw_headers(painter, total_w, total_h, hh, rw, 3, 5)
            painter.end()
            return

        start_col_idx = self._col_to_idx(self.start_col)
        content_rows = 0
        for g in self.groups:
            if self.use_groups:
                content_rows += 1
            content_rows += math.ceil(len(g["images"]) / self.cols) if g["images"] else 0
            if self.use_groups:
                content_rows += 1

        show_cols = max(self.cols + start_col_idx, start_col_idx + self.cols + 1)
        show_rows = max(content_rows + self.start_row, self.start_row + content_rows + 1)
        self._draw_headers(painter, total_w, total_h, hh, rw, show_cols, show_rows)

        grid_w = total_w - rw
        grid_h = total_h - hh
        cell_w = grid_w / show_cols if show_cols else grid_w
        cell_h = grid_h / show_rows if show_rows else grid_h
        aspect = (self.crop_ratio[0] / self.crop_ratio[1]) if self.crop_ratio else 4 / 3

        current_row = self.start_row - 1
        img_num = 0

        for g in self.groups:
            if self.use_groups:
                hx = rw + start_col_idx * cell_w + 2
                hy = hh + current_row * cell_h
                painter.setPen(QColor("#1a1a1a"))
                painter.setFont(QFont("Arial", 7, QFont.Bold))
                painter.drawText(int(hx), int(hy), int(cell_w * self.cols), int(cell_h),
                                 Qt.AlignLeft | Qt.AlignVCenter, g["title"])
                current_row += 1

            img_rows = math.ceil(len(g["images"]) / self.cols) if g["images"] else 0
            for r in range(img_rows):
                for c in range(self.cols):
                    idx = r * self.cols + c
                    if idx >= len(g["images"]):
                        break
                    img_num += 1
                    grid_col = start_col_idx + c
                    grid_row = current_row + r
                    cx = rw + grid_col * cell_w + 1
                    cy = hh + grid_row * cell_h + 1
                    cw = cell_w - 2
                    ch = cell_h - 2
                    img_aspect = aspect
                    cell_aspect = cw / ch if ch > 0 else 1
                    if img_aspect > cell_aspect:
                        iw, ih = cw, cw / img_aspect
                    else:
                        ih, iw = ch, ch * img_aspect
                    ix = cx + (cw - iw) / 2
                    iy = cy + (ch - ih) / 2
                    painter.fillRect(int(ix), int(iy), int(iw), int(ih), QColor("#6366f1"))
                    painter.setPen(QColor("#fff"))
                    painter.setFont(QFont("Arial", 7))
                    painter.drawText(int(ix), int(iy), int(iw), int(ih), Qt.AlignCenter, str(img_num))

            current_row += img_rows + (1 if self.use_groups else 0)

        painter.end()

    def _draw_headers(self, painter, total_w, total_h, hh, rw, show_cols, show_rows):
        grid_w = total_w - rw
        grid_h = total_h - hh
        cell_w = grid_w / show_cols if show_cols else grid_w
        cell_h = grid_h / show_rows if show_rows else grid_h
        painter.fillRect(rw, 0, int(grid_w), hh, QColor("#e0e0e0"))
        painter.fillRect(0, hh, rw, int(grid_h), QColor("#e0e0e0"))
        painter.fillRect(0, 0, rw, hh, QColor("#d0d0d0"))
        painter.setPen(QPen(QColor("#c0c0c0"), 1))
        for c in range(show_cols + 1):
            x = int(rw + c * cell_w)
            painter.drawLine(x, 0, x, total_h)
        for r in range(show_rows + 1):
            y = int(hh + r * cell_h)
            painter.drawLine(0, y, total_w, y)
        painter.setPen(QPen(QColor("#999"), 1))
        painter.drawLine(0, hh, total_w, hh)
        painter.drawLine(rw, 0, rw, total_h)
        painter.setPen(QColor("#333"))
        painter.setFont(QFont("Arial", 7))
        for c in range(show_cols):
            x = int(rw + c * cell_w)
            letter = get_column_letter(c + 1)
            painter.drawText(x, 0, int(cell_w), hh, Qt.AlignCenter, letter)
        for r in range(show_rows):
            y = int(hh + r * cell_h)
            painter.drawText(0, y, rw, int(cell_h), Qt.AlignCenter, str(r + 1))

    @staticmethod
    def _col_to_idx(col_str):
        col_str = col_str.upper().strip()
        idx = 0
        for ch in col_str:
            if ch.isalpha():
                idx = idx * 26 + (ord(ch) - ord('A'))
        return idx


# ── Main window ────────────────────────────────────────────────────────────────
class MainWindow(QMainWindow):
    GROUP_ROLE = Qt.UserRole + 1  # stores group index
    PATH_ROLE = Qt.UserRole + 2  # stores image path
    TYPE_ROLE = Qt.UserRole + 3  # "group" or "image"

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel Image Inserter")
        self.setMinimumSize(750, 750)
        self.groups = [{"title": "Group 1", "images": []}]
        self._collapsed_groups = set()  # indices of collapsed groups
        self._build_ui()

    @property
    def image_paths(self):
        return [p for g in self.groups for p in g["images"]]

    def _build_ui(self):
        central = QWidget()
        self.setCentralWidget(central)
        root = QVBoxLayout(central)
        root.setSpacing(6)

        # ── Excel file ─────────────────────────────────────────────────────
        # Header row: title + about button on same line
        file_header = QHBoxLayout()
        file_header.addWidget(QLabel("<b>Excel File</b>"))
        file_header.addStretch()
        self.btn_about = QPushButton("?")
        self.btn_about.setFixedSize(24, 24)
        self.btn_about.setToolTip("About")
        self.btn_about.setStyleSheet(
            "QPushButton { border: 1px solid palette(mid); border-radius: 12px; "
            "font-weight: bold; font-size: 13px; }"
            "QPushButton:hover { background: palette(midlight); }"
        )
        self.btn_about.clicked.connect(self._show_about)
        file_header.addWidget(self.btn_about)
        root.addLayout(file_header)

        grp_file = QGroupBox()
        lay_file = QVBoxLayout(grp_file)
        lay_file.setSpacing(4)

        lbl_format = QLabel("\u26a0 Only .xlsx (Excel 2007+) is supported. Old .xls files must be re-saved as .xlsx first.")
        lbl_format.setStyleSheet("color: #e67e22; font-size: 11px; padding: 2px 0;")
        lbl_format.setWordWrap(True)
        lay_file.addWidget(lbl_format)

        row1 = QHBoxLayout()
        self.rb_new = QRadioButton("Create new")
        self.rb_open = QRadioButton("Open existing")
        self.rb_open.setChecked(True)
        bg = QButtonGroup(self)
        bg.addButton(self.rb_new)
        bg.addButton(self.rb_open)
        row1.addWidget(self.rb_open)
        row1.addWidget(self.rb_new)
        lay_file.addLayout(row1)

        row2 = QHBoxLayout()
        self.le_file = QLineEdit()
        self.le_file.setPlaceholderText("Path to .xlsx file")
        self.btn_browse_file = QPushButton("Browse...")
        self.btn_browse_file.clicked.connect(self._browse_file)
        row2.addWidget(self.le_file, 1)
        row2.addWidget(self.btn_browse_file)
        lay_file.addLayout(row2)

        row3 = QHBoxLayout()
        row3.addWidget(QLabel("Sheet:"))
        self.combo_sheet = QComboBox()
        self.combo_sheet.setMinimumWidth(120)
        row3.addWidget(self.combo_sheet, 1)
        self.cb_new_sheet = QCheckBox("New:")
        self.le_new_sheet = QLineEdit()
        self.le_new_sheet.setPlaceholderText("Sheet name")
        self.le_new_sheet.setEnabled(False)
        self.cb_new_sheet.toggled.connect(self.le_new_sheet.setEnabled)
        self.cb_new_sheet.toggled.connect(lambda v: self.combo_sheet.setEnabled(not v))
        self.cb_new_sheet.toggled.connect(self._on_new_sheet_toggled)
        row3.addWidget(self.cb_new_sheet)
        row3.addWidget(self.le_new_sheet)
        lay_file.addLayout(row3)

        # Insert after selector (for new sheets)
        row_insert = QHBoxLayout()
        self.lbl_insert_after = QLabel("Insert after:")
        self.combo_insert_after = QComboBox()
        self.combo_insert_after.addItem("(at the end)")
        row_insert.addWidget(self.lbl_insert_after)
        row_insert.addWidget(self.combo_insert_after, 1)
        self.lbl_insert_after.hide()
        self.combo_insert_after.hide()
        lay_file.addLayout(row_insert)

        # TOC checkbox
        self.cb_toc = QCheckBox("Create / update Contents sheet with links")
        self.cb_toc.setChecked(True)
        self.cb_toc.hide()
        lay_file.addWidget(self.cb_toc)

        self.rb_new.toggled.connect(self._on_file_mode_changed)
        root.addWidget(grp_file)

        # ── Images ─────────────────────────────────────────────────────────
        grp_img = QGroupBox("Images")
        lay_img = QVBoxLayout(grp_img)

        # Mode toggle
        mode_row = QHBoxLayout()
        self.cb_use_groups = QCheckBox("Use groups (headers + TOC)")
        self.cb_use_groups.toggled.connect(self._on_group_mode_toggled)
        mode_row.addWidget(self.cb_use_groups)
        mode_row.addStretch()
        lay_img.addLayout(mode_row)

        # Image/group controls
        btn_row = QHBoxLayout()
        self.btn_add_img = QPushButton("Add images...")
        self.btn_add_img.clicked.connect(self._add_images)
        self.btn_add_group = QPushButton("+ Group")
        self.btn_add_group.clicked.connect(self._add_group)
        self.btn_add_group.hide()
        self.btn_remove = QPushButton("Remove")
        self.btn_remove.clicked.connect(self._remove_selected)
        self.btn_clear_img = QPushButton("Clear all")
        self.btn_clear_img.clicked.connect(self._clear_images)
        self.btn_move_up = QPushButton("\u25b2")
        self.btn_move_up.setMaximumWidth(28)
        self.btn_move_up.setToolTip("Move up")
        self.btn_move_up.clicked.connect(lambda: self._move_selected(-1))
        self.btn_move_down = QPushButton("\u25bc")
        self.btn_move_down.setMaximumWidth(28)
        self.btn_move_down.setToolTip("Move down")
        self.btn_move_down.clicked.connect(lambda: self._move_selected(1))
        btn_row.addWidget(self.btn_add_img)
        btn_row.addWidget(self.btn_add_group)
        btn_row.addWidget(self.btn_remove)
        btn_row.addWidget(self.btn_clear_img)
        btn_row.addWidget(self.btn_move_up)
        btn_row.addWidget(self.btn_move_down)
        btn_row.addStretch()

        # View switcher
        self.btn_view_list = QPushButton("List")
        self.btn_view_detail = QPushButton("Details")
        self.btn_view_stack = QPushButton("Stack")
        for b in [self.btn_view_list, self.btn_view_detail, self.btn_view_stack]:
            b.setCheckable(True)
            b.setMaximumWidth(65)
            b.setStyleSheet("QPushButton:checked{background:#6366f1;color:#fff;border-radius:4px;padding:3px 6px}")
        self.btn_view_list.setChecked(True)
        self.btn_view_list.clicked.connect(lambda: self._switch_view("list"))
        self.btn_view_detail.clicked.connect(lambda: self._switch_view("detail"))
        self.btn_view_stack.clicked.connect(lambda: self._switch_view("stack"))
        btn_row.addWidget(self.btn_view_list)
        btn_row.addWidget(self.btn_view_detail)
        btn_row.addWidget(self.btn_view_stack)
        lay_img.addLayout(btn_row)

        # Active group selector
        group_sel_row = QHBoxLayout()
        self.lbl_active_group = QLabel("Add to group:")
        self.lbl_active_group.hide()
        self.combo_active_group = QComboBox()
        self.combo_active_group.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.combo_active_group.hide()
        group_sel_row.addWidget(self.lbl_active_group)
        group_sel_row.addWidget(self.combo_active_group)
        lay_img.addLayout(group_sel_row)

        # View: List (thumbnails + groups)
        self.tree = QTreeWidget()
        self.tree.setHeaderLabels(["", "Name", "Size", "After", ""])
        self.tree.setIconSize(QSize(48, 48))
        self.tree.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.tree.setRootIsDecorated(False)
        self.tree.setColumnWidth(0, 56)
        self.tree.setColumnWidth(1, 200)
        self.tree.setColumnWidth(2, 70)
        self.tree.setColumnWidth(3, 70)
        self.tree.setColumnWidth(4, 30)
        self.tree.header().setStretchLastSection(False)
        self.tree.header().setSectionResizeMode(1, QHeaderView.Stretch)
        self.tree.setMinimumHeight(200)
        self.tree.itemClicked.connect(self._on_tree_click)
        self.tree.setDragDropMode(QAbstractItemView.InternalMove)
        self.tree.setDefaultDropAction(Qt.MoveAction)
        self.tree.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tree.customContextMenuRequested.connect(self._on_tree_context_menu)
        lay_img.addWidget(self.tree)

        # View: Details (no thumbnails)
        self.tree_detail = QTreeWidget()
        self.tree_detail.setHeaderLabels(["", "Name", "Dimensions", "Size", "After", ""])
        self.tree_detail.setSelectionMode(QAbstractItemView.ExtendedSelection)
        self.tree_detail.setRootIsDecorated(False)
        self.tree_detail.setColumnWidth(0, 30)
        self.tree_detail.setColumnWidth(1, 200)
        self.tree_detail.setColumnWidth(2, 90)
        self.tree_detail.setColumnWidth(3, 70)
        self.tree_detail.setColumnWidth(4, 70)
        self.tree_detail.setColumnWidth(5, 30)
        self.tree_detail.header().setStretchLastSection(False)
        self.tree_detail.header().setSectionResizeMode(1, QHeaderView.Stretch)
        self.tree_detail.setMinimumHeight(200)
        self.tree_detail.itemClicked.connect(self._on_tree_detail_click)
        self.tree_detail.setContextMenuPolicy(Qt.CustomContextMenu)
        self.tree_detail.customContextMenuRequested.connect(self._on_tree_context_menu_detail)
        self.tree_detail.hide()
        lay_img.addWidget(self.tree_detail)

        # View: Stack (thumbnail cards)
        self.thumb_stack = ThumbStackView()
        self.thumb_stack.delete_requested.connect(self._delete_by_path_flat)
        self.thumb_stack.order_changed.connect(self._on_stack_reorder)
        self.thumb_stack.setMinimumHeight(200)
        self.thumb_stack.hide()
        lay_img.addWidget(self.thumb_stack)

        # Bottom: separator + count + total (fixed below tree)
        sep_line = QFrame()
        sep_line.setFrameShape(QFrame.HLine)
        sep_line.setFrameShadow(QFrame.Sunken)
        lay_img.addWidget(sep_line)
        bottom_row = QHBoxLayout()
        bottom_row.setContentsMargins(4, 2, 4, 2)
        self.lbl_img_count = QLabel("0 images")
        self.lbl_total_size = QLabel("")
        self.lbl_total_size.setStyleSheet("color:#999")
        bottom_row.addWidget(self.lbl_img_count)
        bottom_row.addStretch()
        bottom_row.addWidget(self.lbl_total_size)
        lay_img.addLayout(bottom_row)

        root.addWidget(grp_img, 1)

        # ── Settings: Resize + Display in one row ──────────────────────────
        settings_row = QHBoxLayout()

        grp_resize = QGroupBox("Resize (px)")
        g_r = QGridLayout(grp_resize)
        g_r.setSpacing(4)
        resize_presets = ["Auto", "64", "128", "256", "320", "480", "640", "800", "1024", "1200", "1600", "1920", "2048", "3840"]
        g_r.addWidget(QLabel("W:"), 0, 0)
        self.combo_px_w = QComboBox()
        self.combo_px_w.setEditable(True)
        self.combo_px_w.addItems(resize_presets)
        self.combo_px_w.setCurrentText("1200")
        self.combo_px_w.currentTextChanged.connect(self._on_resize_changed)
        g_r.addWidget(self.combo_px_w, 0, 1)
        g_r.addWidget(QLabel("H:"), 1, 0)
        self.combo_px_h = QComboBox()
        self.combo_px_h.setEditable(True)
        self.combo_px_h.addItems(resize_presets)
        self.combo_px_h.setCurrentText("Auto")
        self.combo_px_h.currentTextChanged.connect(self._on_resize_changed)
        g_r.addWidget(self.combo_px_h, 1, 1)
        settings_row.addWidget(grp_resize)

        grp_display = QGroupBox("Display (cm)")
        g_d = QGridLayout(grp_display)
        g_d.setSpacing(4)
        g_d.addWidget(QLabel("Mode:"), 0, 0)
        self.combo_display_mode = QComboBox()
        self.combo_display_mode.addItems(["Per image", "Fixed ratio", "Manual"])
        self.combo_display_mode.setCurrentIndex(2)
        self.combo_display_mode.currentIndexChanged.connect(self._on_display_mode_changed)
        g_d.addWidget(self.combo_display_mode, 0, 1)
        g_d.addWidget(QLabel("W:"), 1, 0)
        self.spin_cm_w = QDoubleSpinBox()
        self.spin_cm_w.setRange(0.5, 50)
        self.spin_cm_w.setValue(6.0)
        self.spin_cm_w.setSingleStep(0.5)
        self.spin_cm_w.setSuffix(" cm")
        self.spin_cm_w.valueChanged.connect(self._on_cm_w_changed)
        g_d.addWidget(self.spin_cm_w, 1, 1)
        g_d.addWidget(QLabel("H:"), 2, 0)
        self.spin_cm_h = QDoubleSpinBox()
        self.spin_cm_h.setRange(0.5, 50)
        self.spin_cm_h.setValue(4.5)
        self.spin_cm_h.setSingleStep(0.5)
        self.spin_cm_h.setSuffix(" cm")
        self.spin_cm_h.valueChanged.connect(self._on_cm_h_changed)
        g_d.addWidget(self.spin_cm_h, 2, 1)
        self._display_aspect = 6.0 / 4.5
        self._cm_updating = False
        settings_row.addWidget(grp_display)

        grp_crop = QGroupBox("Crop")
        g_c = QVBoxLayout(grp_crop)
        self.combo_crop = QComboBox()
        self.combo_crop.addItems(CROP_PRESETS.keys())
        self.combo_crop.currentTextChanged.connect(self._on_settings_changed)
        self.combo_crop.currentTextChanged.connect(self._on_resize_changed)
        g_c.addWidget(self.combo_crop)
        g_c.addStretch()
        settings_row.addWidget(grp_crop)

        root.addLayout(settings_row)

        # ── Grid + Position + Preview ──────────────────────────────────────
        grid_row = QHBoxLayout()

        grp_grid = QGroupBox("Grid")
        g_g = QGridLayout(grp_grid)
        g_g.setSpacing(4)
        g_g.addWidget(QLabel("Cols:"), 0, 0)
        self.spin_cols = QSpinBox()
        self.spin_cols.setRange(1, 20)
        self.spin_cols.setValue(2)
        self.spin_cols.valueChanged.connect(self._on_settings_changed)
        g_g.addWidget(self.spin_cols, 0, 1)
        g_g.addWidget(QLabel("H gap:"), 1, 0)
        self.spin_gap_h = QDoubleSpinBox()
        self.spin_gap_h.setRange(0, 50)
        self.spin_gap_h.setValue(0.5)
        self.spin_gap_h.setSingleStep(0.05)
        self.spin_gap_h.setSuffix(" cm")
        self.spin_gap_h.setDecimals(2)
        g_g.addWidget(self.spin_gap_h, 1, 1)
        g_g.addWidget(QLabel("V gap:"), 2, 0)
        self.spin_gap_v = QDoubleSpinBox()
        self.spin_gap_v.setRange(0, 50)
        self.spin_gap_v.setValue(0.5)
        self.spin_gap_v.setSingleStep(0.05)
        self.spin_gap_v.setSuffix(" cm")
        self.spin_gap_v.setDecimals(2)
        g_g.addWidget(self.spin_gap_v, 2, 1)
        grid_row.addWidget(grp_grid)

        grp_pos = QGroupBox("Position")
        g_p = QGridLayout(grp_pos)
        g_p.setSpacing(4)
        g_p.addWidget(QLabel("Cell:"), 0, 0)
        pos_row = QHBoxLayout()
        self.le_start_col = QLineEdit("A")
        self.le_start_col.setMaximumWidth(35)
        self.spin_start_row = QSpinBox()
        self.spin_start_row.setRange(1, 1048576)
        self.spin_start_row.setValue(1)
        pos_row.addWidget(self.le_start_col)
        pos_row.addWidget(self.spin_start_row)
        g_p.addLayout(pos_row, 0, 1)
        g_p.addWidget(QLabel("Mode:"), 1, 0)
        self.combo_placement = QComboBox()
        self.combo_placement.addItems(["Over cells", "In cell"])
        self.combo_placement.currentIndexChanged.connect(self._on_settings_changed)
        self.le_start_col.textChanged.connect(self._on_settings_changed)
        self.spin_start_row.valueChanged.connect(self._on_settings_changed)
        g_p.addWidget(self.combo_placement, 1, 1)
        grid_row.addWidget(grp_pos)

        self.grid_preview = GridPreview()
        grid_row.addWidget(self.grid_preview, 1)

        root.addLayout(grid_row)

        # ── Action ─────────────────────────────────────────────────────────
        self.progress = QProgressBar()
        self.progress.setValue(0)
        self.progress.setMaximumHeight(16)
        root.addWidget(self.progress)

        action_row = QHBoxLayout()
        self.lbl_status = QLabel("Ready")
        action_row.addWidget(self.lbl_status, 1)
        self.btn_insert = QPushButton("  Insert Images  ")
        self.btn_insert.setMinimumHeight(36)
        self.btn_insert.setStyleSheet("font-size:13px;font-weight:bold;background:#6366f1;color:#fff;border-radius:6px;padding:6px 20px")
        self.btn_insert.clicked.connect(self._do_insert)
        action_row.addWidget(self.btn_insert)
        root.addLayout(action_row)

        # (About is shown via ? button in top header)

        self._rebuild_tree()

    # ── File/sheet management ─────────────────────────────────────────────
    def _on_file_mode_changed(self):
        is_open = self.rb_open.isChecked()
        self.btn_browse_file.setText("Browse..." if is_open else "Save as...")
        if not is_open:
            self.combo_sheet.clear()
            self.cb_new_sheet.setChecked(True)

    def _on_new_sheet_toggled(self, checked):
        self.lbl_insert_after.setVisible(checked and self.combo_sheet.count() > 0)
        self.combo_insert_after.setVisible(checked and self.combo_sheet.count() > 0)
        if checked:
            self.combo_insert_after.clear()
            self.combo_insert_after.addItem("(at the end)")
            for i in range(self.combo_sheet.count()):
                self.combo_insert_after.addItem(self.combo_sheet.itemText(i))

    def _browse_file(self):
        if self.rb_open.isChecked():
            path, _ = QFileDialog.getOpenFileName(self, "Open Excel", "", "Excel Files (*.xlsx)")
        else:
            path, _ = QFileDialog.getSaveFileName(self, "Save Excel As", "images.xlsx", "Excel Files (*.xlsx)")
        if path:
            self.le_file.setText(path)
            if self.rb_open.isChecked() and os.path.exists(path):
                try:
                    wb = openpyxl.load_workbook(path, read_only=True)
                    self.combo_sheet.clear()
                    self.combo_sheet.addItems(wb.sheetnames)
                    wb.close()
                except Exception as e:
                    QMessageBox.warning(self, "Error", str(e))

    # ── Group mode toggle ─────────────────────────────────────────────────
    def _on_group_mode_toggled(self, enabled):
        self.btn_add_group.setVisible(enabled)
        self.cb_toc.setVisible(enabled)
        self.lbl_active_group.setVisible(enabled)
        self.combo_active_group.setVisible(enabled)
        if not enabled:
            all_images = self.image_paths
            self.groups = [{"title": "All Images", "images": all_images}]
        self._rebuild_tree()

    # ── Group management ──────────────────────────────────────────────────
    def _add_group(self):
        name, ok = QInputDialog.getText(self, "New Group", "Group title:")
        if ok and name.strip():
            self.groups.append({"title": name.strip(), "images": []})
            self._rebuild_tree()
            # Auto-select newly created group
            self.combo_active_group.setCurrentIndex(len(self.groups) - 1)
            self._on_settings_changed()

    def _get_selected_group_idx(self):
        """Get group index of currently selected item."""
        items = self.tree.selectedItems()
        if not items:
            return len(self.groups) - 1 if self.groups else -1
        item = items[0]
        tp = item.data(0, self.TYPE_ROLE)
        if tp == "group":
            return item.data(0, self.GROUP_ROLE)
        elif tp == "image":
            return item.data(0, self.GROUP_ROLE)
        return 0

    # ── Tree view ─────────────────────────────────────────────────────────
    def _rebuild_tree(self):
        max_w, max_h = self._get_resize_px()
        use_groups = self.cb_use_groups.isChecked()

        # ── Update group selector combo ──
        prev_idx = self.combo_active_group.currentIndex()
        self.combo_active_group.blockSignals(True)
        self.combo_active_group.clear()
        for gi, group in enumerate(self.groups):
            self.combo_active_group.addItem(group["title"], gi)
        if 0 <= prev_idx < len(self.groups):
            self.combo_active_group.setCurrentIndex(prev_idx)
        elif self.groups:
            self.combo_active_group.setCurrentIndex(len(self.groups) - 1)
        self.combo_active_group.blockSignals(False)

        # ── List view ──
        self.tree.clear()
        for gi, group in enumerate(self.groups):
            if use_groups:
                collapsed = gi in self._collapsed_groups
                icon = GROUP_ICON_COLLAPSED if collapsed else GROUP_ICON
                grp_item = QTreeWidgetItem([
                    icon,
                    f"{group['title']} ({len(group['images'])})",
                    "", "", ""
                ])
                grp_item.setData(0, self.TYPE_ROLE, "group")
                grp_item.setData(0, self.GROUP_ROLE, gi)
                grp_font = grp_item.font(1)
                grp_font.setBold(True)
                grp_item.setFont(1, grp_font)
                base = self.palette().color(self.backgroundRole())
                is_dark = base.lightnessF() < 0.5
                if is_dark:
                    grp_bg = QColor(base.red() + (255 - base.red()) // 5,
                                    base.green() + (255 - base.green()) // 5,
                                    base.blue() + (255 - base.blue()) // 5)
                else:
                    grp_bg = QColor(base.red() - base.red() // 10,
                                    base.green() - base.green() // 10,
                                    base.blue() - base.blue() // 10)
                grp_fg = QColor(Qt.white) if is_dark else QColor(Qt.black)
                for c in range(5):
                    grp_item.setBackground(c, grp_bg)
                    grp_item.setForeground(c, grp_fg)
                self.tree.addTopLevelItem(grp_item)
                if collapsed:
                    continue

            for p in group["images"]:
                orig_mb, est_mb, w, h = estimate_size(p, max_w, max_h)
                item = QTreeWidgetItem(["", Path(p).name, f"{orig_mb:.2f} MB", f"{est_mb:.2f} MB", "\u00d7"])
                item.setData(0, self.TYPE_ROLE, "image")
                item.setData(0, self.PATH_ROLE, p)
                item.setData(0, self.GROUP_ROLE, gi)
                try:
                    px = QPixmap(p).scaled(48, 48, Qt.KeepAspectRatio, Qt.SmoothTransformation)
                    item.setIcon(0, QIcon(px))
                except Exception:
                    pass
                self.tree.addTopLevelItem(item)

        # ── Detail view ──
        self.tree_detail.clear()
        for gi, group in enumerate(self.groups):
            if use_groups:
                collapsed = gi in self._collapsed_groups
                icon = GROUP_ICON_COLLAPSED if collapsed else GROUP_ICON
                grp_item = QTreeWidgetItem([icon, f"{group['title']} ({len(group['images'])})", "", "", "", ""])
                grp_item.setData(0, self.TYPE_ROLE, "group")
                grp_item.setData(0, self.GROUP_ROLE, gi)
                grp_font = grp_item.font(1)
                grp_font.setBold(True)
                grp_item.setFont(1, grp_font)
                base = self.palette().color(self.backgroundRole())
                is_dark = base.lightnessF() < 0.5
                if is_dark:
                    grp_bg = QColor(base.red() + (255 - base.red()) // 5,
                                    base.green() + (255 - base.green()) // 5,
                                    base.blue() + (255 - base.blue()) // 5)
                else:
                    grp_bg = QColor(base.red() - base.red() // 10,
                                    base.green() - base.green() // 10,
                                    base.blue() - base.blue() // 10)
                grp_fg = QColor(Qt.white) if is_dark else QColor(Qt.black)
                for c in range(6):
                    grp_item.setBackground(c, grp_bg)
                    grp_item.setForeground(c, grp_fg)
                self.tree_detail.addTopLevelItem(grp_item)
                if collapsed:
                    continue
            for p in group["images"]:
                orig_mb, est_mb, w, h = estimate_size(p, max_w, max_h)
                dim = f"{w}\u00d7{h}" if w else "?"
                item = QTreeWidgetItem(["", Path(p).name, dim, f"{orig_mb:.2f} MB", f"{est_mb:.2f} MB", "\u00d7"])
                item.setData(0, self.TYPE_ROLE, "image")
                item.setData(0, self.PATH_ROLE, p)
                item.setData(0, self.GROUP_ROLE, gi)
                self.tree_detail.addTopLevelItem(item)

        # ── Stack view ──
        all_images = self.image_paths
        self.thumb_stack.set_images(all_images, max_w, max_h)

        self._update_count()

    def _switch_view(self, mode):
        self.btn_view_list.setChecked(mode == "list")
        self.btn_view_detail.setChecked(mode == "detail")
        self.btn_view_stack.setChecked(mode == "stack")
        self.tree.setVisible(mode == "list")
        self.tree_detail.setVisible(mode == "detail")
        self.thumb_stack.setVisible(mode == "stack")

    def _on_tree_detail_click(self, item, col):
        tp = item.data(0, self.TYPE_ROLE)
        if tp == "group":
            gi = item.data(0, self.GROUP_ROLE)
            if col <= 1:
                if gi in self._collapsed_groups:
                    self._collapsed_groups.discard(gi)
                else:
                    self._collapsed_groups.add(gi)
                self._rebuild_tree()
        elif tp == "image" and col == 5:
            path = item.data(0, self.PATH_ROLE)
            if path:
                self._delete_by_path(path, item.data(0, self.GROUP_ROLE))

    def _on_tree_context_menu_detail(self, pos):
        item = self.tree_detail.itemAt(pos)
        if not item:
            return
        tp = item.data(0, self.TYPE_ROLE)
        menu = QMenu(self)
        if tp == "group":
            gi = item.data(0, self.GROUP_ROLE)
            menu.addAction("Rename group", lambda: self._rename_group(gi))
            if len(self.groups) > 1:
                menu.addAction("Delete group", lambda: self._delete_group(gi))
        elif tp == "image":
            path = item.data(0, self.PATH_ROLE)
            gi = item.data(0, self.GROUP_ROLE)
            if self.cb_use_groups.isChecked() and len(self.groups) > 1:
                move_menu = menu.addMenu("Move to group...")
                for i, g in enumerate(self.groups):
                    if i != gi:
                        move_menu.addAction(g["title"], lambda p=path, src=gi, dst=i: self._move_image_to_group(p, src, dst))
            menu.addAction("Remove", lambda: self._delete_by_path(path, gi))
        menu.exec_(self.tree_detail.viewport().mapToGlobal(pos))

    def _delete_by_path_flat(self, path):
        """Delete from stack view — find which group has it."""
        for gi, g in enumerate(self.groups):
            if path in g["images"]:
                self._delete_by_path(path, gi)
                return

    def _on_stack_reorder(self, new_order):
        """Reorder from stack view — applies to first group only in flat mode."""
        if not self.cb_use_groups.isChecked() and len(self.groups) == 1:
            self.groups[0]["images"] = new_order
            self._rebuild_tree()

    def _on_tree_click(self, item, col):
        tp = item.data(0, self.TYPE_ROLE)
        if tp == "group":
            gi = item.data(0, self.GROUP_ROLE)
            if col <= 1:
                # Toggle expand/collapse
                if gi in self._collapsed_groups:
                    self._collapsed_groups.discard(gi)
                else:
                    self._collapsed_groups.add(gi)
                self._rebuild_tree()
            return
        if tp == "image" and col == 4:
            path = item.data(0, self.PATH_ROLE)
            if path:
                self._delete_by_path(path, item.data(0, self.GROUP_ROLE))

    def _on_tree_context_menu(self, pos):
        item = self.tree.itemAt(pos)
        if not item:
            return
        tp = item.data(0, self.TYPE_ROLE)
        menu = QMenu(self)

        if tp == "group":
            gi = item.data(0, self.GROUP_ROLE)
            menu.addAction("Rename group", lambda: self._rename_group(gi))
            if len(self.groups) > 1:
                menu.addAction("Delete group", lambda: self._delete_group(gi))
            if gi > 0:
                menu.addAction("Move group up", lambda: self._move_group(gi, -1))
            if gi < len(self.groups) - 1:
                menu.addAction("Move group down", lambda: self._move_group(gi, 1))
        elif tp == "image":
            path = item.data(0, self.PATH_ROLE)
            gi = item.data(0, self.GROUP_ROLE)
            if self.cb_use_groups.isChecked() and len(self.groups) > 1:
                move_menu = menu.addMenu("Move to group...")
                for i, g in enumerate(self.groups):
                    if i != gi:
                        move_menu.addAction(g["title"], lambda p=path, src=gi, dst=i: self._move_image_to_group(p, src, dst))
            menu.addAction("Remove", lambda: self._delete_by_path(path, gi))

        menu.exec_(self.tree.viewport().mapToGlobal(pos))

    def _rename_group(self, gi):
        g = self.groups[gi]
        name, ok = QInputDialog.getText(self, "Rename Group", "New title:", text=g["title"])
        if ok and name.strip():
            g["title"] = name.strip()
            self._rebuild_tree()
            self._on_settings_changed()

    def _delete_group(self, gi):
        if len(self.groups) <= 1:
            return
        g = self.groups[gi]
        if g["images"]:
            if QMessageBox.question(self, "Delete Group",
                                    f"Delete '{g['title']}' with {len(g['images'])} images?",
                                    QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
                return
        self.groups.pop(gi)
        self._collapsed_groups.discard(gi)
        # Re-index collapsed groups
        self._collapsed_groups = {i - 1 if i > gi else i for i in self._collapsed_groups if i != gi}
        self._rebuild_tree()
        self._on_settings_changed()

    def _move_group(self, gi, direction):
        new_gi = gi + direction
        if new_gi < 0 or new_gi >= len(self.groups):
            return
        self.groups[gi], self.groups[new_gi] = self.groups[new_gi], self.groups[gi]
        # Update collapsed set
        new_collapsed = set()
        for c in self._collapsed_groups:
            if c == gi:
                new_collapsed.add(new_gi)
            elif c == new_gi:
                new_collapsed.add(gi)
            else:
                new_collapsed.add(c)
        self._collapsed_groups = new_collapsed
        self._rebuild_tree()
        self._on_settings_changed()

    def _move_image_to_group(self, path, src_gi, dst_gi):
        if path in self.groups[src_gi]["images"]:
            self.groups[src_gi]["images"].remove(path)
            self.groups[dst_gi]["images"].append(path)
            self._rebuild_tree()

    # ── Image management ──────────────────────────────────────────────────
    def _add_images(self):
        if self.cb_use_groups.isChecked():
            gi = self.combo_active_group.currentData()
            if gi is None:
                gi = 0
        else:
            gi = 0
        paths, _ = QFileDialog.getOpenFileNames(
            self, "Select Images", "",
            "Images (*.jpg *.jpeg *.png *.bmp *.webp *.tiff);;All Files (*)"
        )
        new_paths = [p for p in paths if p not in self.groups[gi]["images"]]
        if not new_paths:
            return
        self.groups[gi]["images"].extend(new_paths)
        # Make sure this group is expanded
        self._collapsed_groups.discard(gi)
        self._rebuild_tree()

    def _clear_images(self):
        total = sum(len(g["images"]) for g in self.groups)
        if total == 0:
            return
        if QMessageBox.question(self, "Clear", f"Remove all {total} images?",
                                QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return
        for g in self.groups:
            g["images"].clear()
        self._rebuild_tree()

    def _remove_selected(self):
        items = self.tree.selectedItems()
        if not items:
            return
        to_remove = []  # (gi, path) pairs
        groups_to_remove = []
        for item in items:
            tp = item.data(0, self.TYPE_ROLE)
            if tp == "image":
                to_remove.append((item.data(0, self.GROUP_ROLE), item.data(0, self.PATH_ROLE)))
            elif tp == "group" and self.cb_use_groups.isChecked():
                groups_to_remove.append(item.data(0, self.GROUP_ROLE))

        if not to_remove and not groups_to_remove:
            return

        desc = f"{len(to_remove)} image(s)" if to_remove else ""
        if groups_to_remove:
            desc += f"{', ' if desc else ''}{len(groups_to_remove)} group(s)"
        if QMessageBox.question(self, "Remove", f"Remove {desc}?",
                                QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return

        for gi, path in to_remove:
            if gi < len(self.groups) and path in self.groups[gi]["images"]:
                self.groups[gi]["images"].remove(path)

        for gi in sorted(groups_to_remove, reverse=True):
            if len(self.groups) > 1:
                self.groups.pop(gi)

        self._rebuild_tree()

    def _delete_by_path(self, path, gi):
        if QMessageBox.question(self, "Remove", f"Remove {Path(path).name}?",
                                QMessageBox.Yes | QMessageBox.No) != QMessageBox.Yes:
            return
        if gi < len(self.groups) and path in self.groups[gi]["images"]:
            self.groups[gi]["images"].remove(path)
        self._rebuild_tree()

    def _move_selected(self, direction):
        items = self.tree.selectedItems()
        if not items:
            return
        item = items[0]
        tp = item.data(0, self.TYPE_ROLE)

        if tp == "group" and self.cb_use_groups.isChecked():
            gi = item.data(0, self.GROUP_ROLE)
            self._move_group(gi, direction)
            return

        if tp == "image":
            gi = item.data(0, self.GROUP_ROLE)
            path = item.data(0, self.PATH_ROLE)
            if gi >= len(self.groups):
                return
            images = self.groups[gi]["images"]
            idx = images.index(path) if path in images else -1
            if idx < 0:
                return
            new_idx = idx + direction
            if new_idx < 0 or new_idx >= len(images):
                return
            images[idx], images[new_idx] = images[new_idx], images[idx]
            # Fast swap in tree
            tree_idx = self.tree.indexOfTopLevelItem(item)
            swap_idx = tree_idx + direction
            if 0 <= swap_idx < self.tree.topLevelItemCount():
                swap_item = self.tree.topLevelItem(swap_idx)
                if swap_item.data(0, self.TYPE_ROLE) == "image" and swap_item.data(0, self.GROUP_ROLE) == gi:
                    self.tree.blockSignals(True)
                    a = self.tree.takeTopLevelItem(max(tree_idx, swap_idx))
                    b = self.tree.takeTopLevelItem(min(tree_idx, swap_idx))
                    self.tree.insertTopLevelItem(min(tree_idx, swap_idx), a)
                    self.tree.insertTopLevelItem(max(tree_idx, swap_idx), b)
                    for i in range(self.tree.topLevelItemCount()):
                        if self.tree.topLevelItem(i).data(0, self.PATH_ROLE) == path:
                            self.tree.setCurrentItem(self.tree.topLevelItem(i))
                            break
                    self.tree.blockSignals(False)
                    return
            self._rebuild_tree()

    # ── Counts and settings ───────────────────────────────────────────────
    def _update_count(self):
        all_images = self.image_paths
        n = len(all_images)
        ng = len(self.groups)
        if self.cb_use_groups.isChecked():
            self.lbl_img_count.setText(f"{n} images in {ng} groups")
        else:
            self.lbl_img_count.setText(f"{n} image{'s' if n != 1 else ''}")
        max_w, max_h = self._get_resize_px()
        total_orig = sum(os.path.getsize(p) / (1024 * 1024) for p in all_images if os.path.exists(p))
        total_est = sum(estimate_size(p, max_w, max_h)[1] for p in all_images)
        if total_orig > 0:
            self.lbl_total_size.setText(f"Total: {total_orig:.2f} MB \u2192 {total_est:.2f} MB")
        else:
            self.lbl_total_size.setText("")
        self._on_settings_changed()

    def _get_resize_px(self):
        def _parse(combo):
            txt = combo.currentText().strip()
            if not txt or txt.lower() == "auto":
                return None
            try:
                return int(txt)
            except ValueError:
                return None
        return _parse(self.combo_px_w), _parse(self.combo_px_h)

    def _on_resize_changed(self, *_):
        if self.image_paths:
            self._rebuild_tree()

    def _show_about(self):
        QMessageBox.about(
            self, "About Excel Image Inserter",
            f"<h3>Excel Image Inserter</h3>"
            f"<p>Version {APP_VERSION} (build {BUILD_NUMBER})</p>"
            f"<p>Created by I.Moskvin using Claude Opus 4.6</p>"
            f"<p>Batch insert images into Excel .xlsx files<br>"
            f"with grouping, TOC, and layout control.</p>"
        )

    def _on_settings_changed(self, *_):
        crop_key = self.combo_crop.currentText()
        crop = CROP_PRESETS.get(crop_key)
        start_col = self.le_start_col.text().strip().upper() or "A"
        self.grid_preview.update_params(
            groups=self.groups,
            cols=self.spin_cols.value(),
            crop_ratio=crop,
            start_col=start_col,
            start_row=self.spin_start_row.value(),
            placement="in_cell" if self.combo_placement.currentIndex() == 1 else "over",
            use_groups=self.cb_use_groups.isChecked(),
        )

    def _on_display_mode_changed(self, index):
        if index == 0:
            self.spin_cm_w.setEnabled(True)
            self.spin_cm_h.setEnabled(False)
        elif index == 1:
            self.spin_cm_w.setEnabled(True)
            self.spin_cm_h.setEnabled(False)
            self._display_aspect = self.spin_cm_w.value() / max(self.spin_cm_h.value(), 0.1)
        else:
            self.spin_cm_w.setEnabled(True)
            self.spin_cm_h.setEnabled(True)

    def _on_cm_w_changed(self, val):
        if self._cm_updating:
            return
        if self.combo_display_mode.currentIndex() == 1:
            self._cm_updating = True
            self.spin_cm_h.setValue(val / max(self._display_aspect, 0.01))
            self._cm_updating = False

    def _on_cm_h_changed(self, val):
        if self._cm_updating:
            return
        if self.combo_display_mode.currentIndex() == 1:
            self._cm_updating = True
            self.spin_cm_w.setValue(val * self._display_aspect)
            self._cm_updating = False

    # ── Insert ────────────────────────────────────────────────────────────
    def _do_insert(self):
        file_path = self.le_file.text().strip()
        if self.rb_open.isChecked() and (not file_path or not os.path.exists(file_path)):
            QMessageBox.warning(self, "Error", "Please select an existing Excel file.")
            return
        if self.rb_new.isChecked() and not file_path:
            QMessageBox.warning(self, "Error", "Please specify a file path to save.")
            return
        if not self.image_paths:
            QMessageBox.warning(self, "Error", "No images to insert.")
            return

        sheet_new = self.cb_new_sheet.isChecked()
        if sheet_new:
            sheet_name = self.le_new_sheet.text().strip()
            if not sheet_name:
                QMessageBox.warning(self, "Error", "Enter a sheet name.")
                return
        else:
            sheet_name = self.combo_sheet.currentText()
            if not sheet_name:
                QMessageBox.warning(self, "Error", "Select a sheet.")
                return

        start_col = self.le_start_col.text().strip().upper()
        if not start_col or not start_col.isalpha():
            QMessageBox.warning(self, "Error", "Column must be a letter (A, B, C...).")
            return

        crop = CROP_PRESETS.get(self.combo_crop.currentText())

        # Determine insert position
        insert_after_name = None
        if sheet_new and self.combo_insert_after.isVisible():
            sel = self.combo_insert_after.currentIndex()
            if sel > 0:
                insert_after_name = self.combo_insert_after.currentText()

        params = {
            "excel_path": file_path if self.rb_open.isChecked() else None,
            "save_path": file_path,
            "sheet_new": sheet_new,
            "sheet_name": sheet_name,
            "insert_after_name": insert_after_name,
            "groups": [dict(g) for g in self.groups],
            "resize_px_w": self._get_resize_px()[0],
            "resize_px_h": self._get_resize_px()[1],
            "display_w_cm": self.spin_cm_w.value(),
            "display_h_cm": self.spin_cm_h.value(),
            "display_mode": self.combo_display_mode.currentIndex(),
            "crop_ratio": crop,
            "grid_cols": self.spin_cols.value(),
            "start_col": start_col,
            "start_row": self.spin_start_row.value(),
            "placement": "in_cell" if self.combo_placement.currentIndex() == 1 else "over",
            "gap_h_cm": self.spin_gap_h.value(),
            "gap_v_cm": self.spin_gap_v.value(),
            "create_toc": self.cb_toc.isChecked() and self.cb_use_groups.isChecked(),
            "use_groups": self.cb_use_groups.isChecked(),
        }

        self.btn_insert.setEnabled(False)
        self.progress.setValue(0)
        self.worker = InsertWorker(params)
        self.worker.progress.connect(self.progress.setValue)
        self.worker.status.connect(self.lbl_status.setText)
        self.worker.finished.connect(self._on_finished)
        self.worker.start()

    def _on_finished(self, error):
        self.btn_insert.setEnabled(True)
        if error:
            self.lbl_status.setText(f"Error: {error}")
            QMessageBox.critical(self, "Error", error)
        else:
            self.progress.setValue(100)
            self.lbl_status.setText("Done!")
            QMessageBox.information(self, "Success", "Images inserted successfully!")


def main():
    app = QApplication(sys.argv)
    app.setStyle("Fusion")
    win = MainWindow()
    screen = app.primaryScreen().availableGeometry()
    win.resize(win.minimumSizeHint().width(), int(screen.height() * 0.9))
    win.move(screen.x(), screen.y())
    win.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()
